__author__ = 'Rays'

import LCCfg
import FeatureMapping

import openpyxl
import codecs
import LCUtil
import os
import pymysql
import random
import sys


drop_table_sql = u'DROP TABLE %s CASCADE; \n'
create_table_sql = u'CREATE TABLE %s ( %s ); \n'
config = LCCfg.LCCfg("default.cfg")
TRAINING_DATA_RATE = 0.10


def get_dict_sheet(fn):
    wb = openpyxl.load_workbook(fn, True)
    ws = wb.get_sheet_by_name(config.dictionary_dictsheet)
    return ws


def get_column_def(ws):
    row_list = []
    for i in range(2, ws.get_highest_row() - 1):
        col_name = ws.cell(row=i, column=1).value.strip()
        desc = ws.cell(row=i, column=2).value.strip().replace('\n', ' ').replace('\'', '\'\'')
        if col_name is not None and desc is not None:
            data_type = ws.cell(row=i, column=3).value.strip().upper()
            row_list.append((col_name, desc, data_type))
    return row_list


def get_data_file_list(dn):
    return [(dn+"/"+fn) for fn in os.listdir(dn) if fn.startswith(config.data_fileprefix) and fn.endswith(".csv")]


def create_lc_data(dn, def_fn, tn):
    str_sql = drop_table_sql % tn
    # col_def_list = get_column_def(get_dict_sheet(def_fn))
    # col_name_list = []
    # for p in col_def_list:
    #     col_name, data_type = p[0], p[2]
    #     if data_type == u"VARCHAR":
    #         col_name_list.append(col_name)
    name_type_dict = get_col_name_type_dict(def_fn)
    max_col_length_dict = raw_data_size(dn, [ name for name, type in name_type_dict.items() if type == u"VARCHAR"])
    col_sql = list()
    for p in name_type_dict.keys():
        data_type = name_type_dict[p]
        if data_type == u"VARCHAR":
            data_type = "VARCHAR(%s)" % str(max_col_length_dict[p])
        col_sql.append("%s %s" % (p, data_type))
    col_sql.append("PRIMARY KEY(id)")
    str_sql += create_table_sql % (tn, ",\n    ".join(col_sql))
    f = codecs.open('sql/CREATE_LC_RAW_DATA.sql', mode='w', encoding='utf-8')
    f.write(str_sql)
    f.close()


def raw_data_size(dn, varchar_list):
    col_pos_dict = {}
    max_col_length_dict = {field:5 for field in varchar_list}
    for fn in get_data_file_list(dn):
        with open(fn) as f:
            for line in f:
                if line.startswith("\"id\""):
                    headers = FeatureMapping.split_raw_data(line)
                    col_pos_dict = FeatureMapping.extract_col_pos_dict(headers)
                elif line.startswith("\""):
                    fields = FeatureMapping.split_raw_data(line)
                    if len(fields) > 0:
                        for v in varchar_list:
                            if v in col_pos_dict:
                                max_col_length_dict[v] = max_col_length_dict[v] \
                                    if max_col_length_dict[v] >= len(fields[col_pos_dict[v]]) \
                                    else len(fields[col_pos_dict[v]])
    return max_col_length_dict


def create_col_def(fn, tn):
    max_col_len = 0
    max_desc_len = 0
    col_def_list = get_column_def(get_dict_sheet(fn))
    for p in col_def_list:
        max_col_len = len(p[0]) if max_col_len < len(p[0]) else max_col_len
        max_desc_len = len(p[1]) if max_desc_len < len(p[1]) else max_desc_len
    str_sql = drop_table_sql % tn
    col_sql = u'NAME VARCHAR(%d) NOT NULL, DESCRIPTION VARCHAR(%d) NOT NULL, PRIMARY KEY(NAME)' % (max_col_len, max_desc_len)
    str_sql += create_table_sql % (tn, col_sql)
    str_sql += '\n'.join(u'INSERT INTO %s VALUES (\'%s\',\'%s\'); ' % (tn, p[0], p[1]) for p in col_def_list)
    f = codecs.open('sql/CREATE_COL_DEF.sql', mode='w', encoding='utf-8')
    f.write(str_sql)
    f.close()


def get_col_name_type_dict(def_fn):
    col_def_list = get_column_def(get_dict_sheet(def_fn))
    col_name_type_dict = {}
    for p in col_def_list:
        col_name_type_dict[p[0]] = p[2]
    return col_name_type_dict


def load_data(dn, def_fn):
    # load column name and type:
    get_col_name_type_dict(def_fn)
    sql = "select count(*) from lc_raw_data"
    conn = pymysql.connect(host=config.appdb_url,
                           port=config.appdb_port,
                           user=config.appdb_user,
                           passwd=config.appdb_password,
                           db=config.appdb_instance)
    cur = conn.cursor()
    cur.execute(sql)
    print(cur.description)
    for row in cur:
        print(row)
    cur.close()
    conn.commit()
    conn.close()


def random_method(line, col_pos_dict):
    return random.random() <= TRAINING_DATA_RATE


def closed_investment(line, col_pos_dict):
    return line[col_pos_dict["loan_status"]] in ["Charged Off", "Default", "Fully Paid", "Late (31-120 days)"]


def random_closed_investment(line, col_pos_dict):
    return closed_investment(line, col_pos_dict) and random_method(line, col_pos_dict)


def sampling(dn, sample_file, sample_method):
    sample_list = []
    col_pos_dict = {}
    for fn in get_data_file_list(dn):
        with open(fn) as f:
            for line in f:
                if line.startswith("\"id\""):
                    if len(sample_list) == 0:
                        sample_list.append(line)
                        headers = FeatureMapping.split_raw_data(line)
                        col_pos_dict = FeatureMapping.extract_col_pos_dict(headers)
                elif line.startswith("\""):
                    content = FeatureMapping.split_raw_data(line)
                    if sample_method(content, col_pos_dict):
                        sample_list.append(line)
        print("Finished sampling %s" % fn)
    f = open(sample_file, "w")
    f.writelines(sample_list)
    f.close()
    print "Sampled size: %d" % len(sample_list)


if len(sys.argv) < 2:
    print 'no argument'
    sys.exit()
else:
    for i in range(1, len(sys.argv)):
        option = sys.argv[i]
        if option.startswith("sql"):
            create_col_def(config.dictionary_file, "col_def")
            create_lc_data(config.data_dir, config.dictionary_file, "lc_raw_data")
        elif option.startswith("data_gen"):
            # generate sample training data
            sample_data_uri = config.data_dir + "/" + config.training_sample
            sampling(config.data_dir, sample_data_uri + ".csv", random_closed_investment)
            x, y = FeatureMapping.map_features(sample_data_uri + ".csv")
            LCUtil.save_mapped_feature(x, y, sample_data_uri)
            # generate full training data
            training_data_uri = config.data_dir + "/" + config.training_full
            sampling(config.data_dir, training_data_uri + ".csv", closed_investment)
            x, y = FeatureMapping.map_features(training_data_uri + ".csv")
            LCUtil.save_mapped_feature(x, y, training_data_uri)
        elif option.startswith("seperate_data_file"):
            # use pre-saved random seeds to ensure the same train/cv/test set
            random_seeds = LCUtil.load_random_seeds()
            sample_data_uri = config.data_dir + "/" + config.training_sample + ".csv"
            LCUtil.separate_training_data_file(sample_data_uri, 0.6, 0.2, random_seeds)
        elif option.startswith("random_seeds"):
            random_seeds = [random.random() for _ in range(20000)]
            LCUtil.save_random_seeds(random_seeds)


